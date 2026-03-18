import os
import shutil
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import re
import numpy as np


def match_image_type_formatting(img_type:str)->str:
    if img_type in ["ph1","dic"]:
        return img_type
    elif img_type == "Phase contrast":
        return "ph1"
    elif img_type == "Differential interference contrast":
        return "dic"
    else:
        raise ValueError(f"Unsupported image type: {img_type}")


# DATA EXTRACTION AND VISUALIZATION FUNCTIONS #
def extract_data(path_input_fn: str, basename_fn: str, image_type: str, interval_in: int, df_assignments_in) -> (dict,list) :
    if not os.path.exists( os.path.join(path_input_fn, basename_fn )):
        print(f"Folder {basename_fn} does not exist. Skipping data extraction...")
        return

    image_type = match_image_type_formatting(image_type)

    folder_path_list = sorted(os.scandir(os.path.join(path_input_fn, basename_fn)), key=lambda x: x.name)
    folder_path_list = [n1 for n1 in folder_path_list if os.path.isdir(n1)]

    metrics = []
    folder_ind=0
    while metrics == [] and folder_ind < len(folder_path_list):
        # frames = len(os.listdir(os.path.join(folder_path_list[folder_ind].path, image_type + "_images")))
        frames = len(np.loadtxt(os.path.join(folder_path_list[folder_ind].path, "segment_"+image_type,"is_broken_vs_frame.txt")))
        print(f"new frames = {frames}")
        metrics = [f for f in os.listdir(os.path.join(folder_path_list[folder_ind].path, "segment_" + image_type)) if f.endswith(".txt")]
        metrics_pillars = [f for f in os.listdir(os.path.join(folder_path_list[folder_ind].path, "track_pillars_" + image_type)) if f.endswith(".txt")]
        tlist = [T * interval_in for T in range(0, frames)]
        folder_ind+=1

    dfs = {}

    if not os.path.exists(os.path.join(path_input_fn, 'code_output_' + basename_fn + '.xlsx')):
        workbook = openpyxl.Workbook()
        workbook.save(os.path.join(path_input_fn, 'code_output_' + basename_fn + '.xlsx'))

    for metric in metrics:
        print(f"\tExtracting data for {metric.split('_vs_')[0]}...")
        dfs[metric.split('_vs_')[0]] = pd.DataFrame({'Frame': range(1, frames + 1), 'Time': tlist})

        for file in folder_path_list:
            try:
                dfs[metric.split('_vs_')[0]][file.name] = pd.read_table(os.path.join(file.path, 'segment_' + image_type, metric), header=None)
            except Exception as e:
                print(e)

        append_to_excel(os.path.join(path_input_fn, 'code_output_' + basename_fn + '.xlsx'), dfs[metric.split('_vs_')[0]], metric.split('_vs_')[0])

    for mp in metrics_pillars:
        header_name = mp.split('.txt')[0]
        excel_output_path = os.path.join(path_input_fn, 'code_output_' + basename_fn + '.xlsx')
        print(f"\tExtracting data for {header_name}...")

        if header_name == "pillar_tracker_x":  # single pass triggered on x; y is read here too
            rows = []
            for file in folder_path_list:
                mp_x = mp
                mp_y = mp.replace("pillar_tracker_x", "pillar_tracker_y")
                try:
                    df_x = pd.read_table(os.path.join(file.path, 'track_pillars_' + image_type, mp_x), header=None)
                    df_y = pd.read_table(os.path.join(file.path, 'track_pillars_' + image_type, mp_y), header=None)

                    for frame_idx in range(frames):
                        x_vals = str(df_x.iloc[frame_idx, 0]).split()
                        y_vals = str(df_y.iloc[frame_idx, 0]).split()
                        n_pillars = min(len(x_vals), len(y_vals))
                        if len(x_vals) != len(y_vals):
                            print(f"Pillar count mismatch for {file.name}, frame {frame_idx + 1}: "
                                f"{len(x_vals)} x-values vs {len(y_vals)} y-values — using first {n_pillars}.")
                        for pillar in range(n_pillars):
                            try:
                                rows.append({
                                    'Sample': file.name,
                                    'Frame': frame_idx + 1,
                                    'Time': tlist[frame_idx],
                                    'Pillar': pillar,
                                    'X': float(x_vals[pillar]),
                                    'Y': float(y_vals[pillar]),
                                })
                            except ValueError:
                                print(f"Invalid value for {file.name}, frame {frame_idx + 1}, pillar {pillar} — skipping.")
                except Exception as e:
                    print(e)

            dfs["pillar_positions"] = pd.DataFrame(rows, columns=['Sample', 'Frame', 'Time', 'Pillar', 'X', 'Y'])
            append_to_excel(excel_output_path, dfs["pillar_positions"], "pillar_positions", 0, 'replace')

            notes_list = [
                "This sheet is in tidy/long format: one row per pillar per frame per sample.",
                "Pillar IDs: 0 = top left, 1 = top right, 2 = bottom right, 3 = bottom left. Bugs may occasionally affect this pattern.",
                f"To verify pillar positions for each sample, check 'pillar_positions.png' in the 'track_pillars_{image_type}' folder."
            ]
            add_notes_to_excel_by_rows(excel_output_path, notes_list, "pillar_positions")

        elif header_name == "pillar_tracker_y":
            pass  # handled in the x pass above

        elif header_name == "pillar_disps_actual" or header_name=="avg_pillar_disps_actual":
            dfs[header_name] = pd.DataFrame({'Frame': range(1, frames + 1), 'Time': tlist})
            for file in folder_path_list:
                try:
                    dfs[header_name][file.name] = pd.read_table(os.path.join(file.path, 'track_pillars_' + image_type, mp), header=None)
                except Exception as e:
                    print(e)
            
            append_to_excel(excel_output_path, dfs[header_name], header_name)
        
        elif header_name == "disp_pillar_dist_to_centroid":
            sheet_name = "disp_pillar_dist_to_centroid"
            rows = []
            for file in folder_path_list:
                try:
                    df = pd.read_table(os.path.join(file.path, 'track_pillars_' + image_type, mp), header=None)

                    for frame_idx in range(frames):
                        vals = str(df.iloc[frame_idx, 0]).split()
                        for pillar in range(len(vals)):
                            try:
                                rows.append({
                                    'Sample': file.name,
                                    'Frame': frame_idx + 1,
                                    'Time': tlist[frame_idx],
                                    'Pillar': pillar,
                                    'Displacement': float(vals[pillar]),
                                })
                            except ValueError:
                                print(f"Invalid value for {file.name}, frame {frame_idx + 1}, pillar {pillar} — skipping.")
                except Exception as e:
                    print(e)

            dfs[sheet_name] = pd.DataFrame(rows, columns=['Sample', 'Frame', 'Time', 'Pillar', 'Displacement'])
            append_to_excel(excel_output_path, dfs[sheet_name], sheet_name, 0, 'replace')

            notes_list = [
                "This sheet is in tidy/long format: one row per pillar per frame per sample.",
                "Each 'Displacement' value is the displacement of a pillar's distance to the centroid of all pillars.",
                "Pillar IDs: 0 = top left, 1 = top right, 2 = bottom right, 3 = bottom left. Bugs may occasionally affect this pattern.",
                f"To verify pillar positions for each sample, check 'pillar_positions.png' in the 'track_pillars_{image_type}' folder."
            ]
            add_notes_to_excel_by_rows(excel_output_path, notes_list, sheet_name)


        # elif header_name == "relative_pillar_distances_pair_names":
        #     pair_name_list = []
        #     for file in folder_path_list:
        #         try:
        #             pillar_pairs = np.loadtxt(
        #                 os.path.join(file.path, 'track_pillars_' + image_type, mp), dtype=str, comments=None
        #                 )
        #             pillar_pairs_string = ', '.join(pillar_pairs)  # Convert to comma-separated string
        #             pair_name_list.append(pillar_pairs_string)
        #         except Exception as e:
        #             print(e)

        # else:
        #     if header_name == "relative_pillar_distances_smoothed_GPR":
        #         header_name = "rel_pillar_dist_GPR"
        #     dfs[header_name] = pd.DataFrame({'Frame': range(1, frames + 1), 'Time': tlist})

        #     for file in folder_path_list:
        #         try:
        #             dfs[header_name][file.name] = pd.read_table(os.path.join(file.path, 'track_pillars_' + image_type, mp), header=None)
        #         except Exception as e:
        #             print(e)

        #     append_to_excel(excel_output_path, dfs[header_name], header_name)

    # if pair_name_list:
    #     notes_list=["This row contains the pillar pair names, corresponding to the values in the cells below. 'p0-p1' means the relative distance between pillar 0 and pillar 1."]
    #     add_notes_to_excel_by_rows(excel_output_path, notes_list, "relative_pillar_distances")
    #     add_notes_to_excel_by_rows(excel_output_path, notes_list, "rel_pillar_dist_GPR")
    #     for pair_ind,pair_name_str in enumerate(pair_name_list):
    #         add_note_to_excel_by_cell(
    #             excel_output_path, pair_name_str, "relative_pillar_distances", excel_row=1, excel_column=4+pair_ind
    #             )
    #         add_note_to_excel_by_cell(
    #             excel_output_path, pair_name_str, "rel_pillar_dist_GPR", excel_row=1, excel_column=4+pair_ind
    #             )
    return

def append_to_excel(fpath, df, sheet_name, start_row_ind=0,sheet_exists_mode='replace'):
    if not os.path.exists(fpath):
        workbook = openpyxl.Workbook()
        workbook.save(fpath)
    with pd.ExcelWriter(fpath, mode="a", if_sheet_exists=sheet_exists_mode) as f:
        df.to_excel(f, sheet_name=sheet_name,startrow=start_row_ind)

def add_notes_to_excel_by_rows(fpath, notes_list, sheet_name):

    wb = openpyxl.load_workbook(fpath)
    ws = wb[sheet_name]

    num_notes = len(notes_list)
    
    # Insert rows at the top for the notes
    ws.insert_rows(1, amount=num_notes)
    
    # Add the note lines
    for j, line in enumerate(notes_list):
        ws.cell(row=j+1, column=1, value=line)

    # Save the modified workbook
    wb.save(fpath)
    return

def add_note_to_excel_by_cell(fpath, note, sheet_name, excel_row, excel_column):

    wb = openpyxl.load_workbook(fpath)
    ws = wb[sheet_name]

    # Add the note line
    ws.cell(row=excel_row, column=excel_column, value=note)

    # Save the modified workbook
    wb.save(fpath)
    return

def conglomerate_segmentation_images(path_output, basename, image_type):
    """
    Finds and copy contour images to a specific folder.

    Parameters:
    - path_output: str, path to the output directory
    - basename: str, base name for the output files
    - image_type: str, type of images (e.g., 'ph1')
    """
    image_type = match_image_type_formatting(image_type)
    # Define the source and destination directories
    if not os.path.exists(os.path.join(path_output, basename)):
        print(f"Folder {basename} does not exist. Skipping visualization...")
        return
    folder_path_list = sorted(os.scandir(os.path.join(path_output, basename)), key=lambda x: x.name)
    folder_path_list = [n1 for n1 in folder_path_list if os.path.isdir(n1)]
    sample_vis_paths = []
    img_paths = []
    for folder in folder_path_list:
        fp = os.path.join(folder.path, "segment_" + image_type, "visualizations")
        images = [f for f in os.listdir(fp) if f.startswith(image_type + "_contour_all_")]
        if images:
            sample_vis_paths.append(fp)
            img_paths+=images

    parent_dir = os.path.join(path_output, "all_samples_segmentation_results")
    if not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)

    basename_dir = os.path.join(path_output, "all_samples_segmentation_results", basename)
    if not os.path.exists(basename_dir):
        os.makedirs(basename_dir, exist_ok=True)
    
    for img_ind,image in enumerate(img_paths):
        src_path = sample_vis_paths[img_ind]
        src_path = os.path.join(src_path, image)
        dest_path = os.path.join(basename_dir, image)
        if not os.path.exists(dest_path):
            shutil.copy2(src_path, dest_path)


def conglomerate_pillar_disps_images(path_output, basename, image_type):
    """
    Finds and copy pillar disps images to a specific folder.

    Parameters:
    - path_output: str, path to the output directory
    - basename: str, base name for the output files
    - image_type: str, type of images (e.g., 'ph1')
    """
    image_type = match_image_type_formatting(image_type)
    # Define the source and destination directories
    if not os.path.exists(os.path.join(path_output, basename)):
        print(f"Folder {basename} does not exist. Skipping visualization...")
        return
    folder_path_list = sorted(os.scandir(os.path.join(path_output, basename)), key=lambda x: x.name)
    folder_path_list = [n1 for n1 in folder_path_list if os.path.isdir(n1)]
    sample_vis_paths = []
    img_paths = []
    for folder in folder_path_list:
        fp = os.path.join(folder.path, "track_pillars_" + image_type)
        images = [f for f in os.listdir(fp) if f.startswith("pillar_disps_and_pillar_contours")]
        if images:
            sample_vis_paths.append(fp)
            img_paths+=images
    # print(img_paths)
    parent_dir = os.path.join(path_output, "all_samples_pillar_tracking_results")
    if not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)

    basename_dir = os.path.join(path_output, "all_samples_pillar_tracking_results", basename)
    if not os.path.exists(basename_dir):
        os.makedirs(basename_dir, exist_ok=True)
    
    for img_ind,image in enumerate(img_paths):
        src_path = sample_vis_paths[img_ind]
        src_path = os.path.join(src_path, image)
        dest_path = os.path.join(basename_dir, image)
        if not os.path.exists(dest_path):
            shutil.copy2(src_path, dest_path)


def visualize_data(path_output_in, basename_in, image_type_in, all_data_in, metrics_in, positions_in, assigned_df_in):
    """
       Visualizes data from all_data_in by plotting metrics over time per condition.

       Parameters:
       - path_output_in: str, path to save output images
       - basename_in: str, base name for output files
       - image_type_in: str, image file type (e.g., 'png', 'jpg')
       - all_data_in: dict, contains dataframes for each metric
       - metrics_in: list, metrics to be plotted
       - positions_in: list, positions to be considered
       - assigned_df_in: DataFrame, contains condition assignments for positions
       """
    # Create a folder to store the visualizations
    if not os.path.exists(os.path.join(path_output_in, basename_in + '_visualizations')):
        os.makedirs(os.path.join(path_output_in, basename_in + '_visualizations'))

    # Create a folder to store the visualizations
    if not os.path.exists(os.path.join(path_output_in, basename_in + '_visualizations', 'segmentation')):
        os.makedirs(os.path.join(path_output_in, basename_in + '_visualizations', 'segmentation'))

    for metric in metrics_in:
        metric = metric.split('_vs_')[0]
        df = all_data_in[metric]
        time_hours = df['Time']  # Convert time to hours

        # Identify position columns
        position_columns = [col for col in df.columns if any(pos in col for pos in positions_in)]

        # Group data by condition
        grouped_data = {}
        # Create a folder to store the visualizations
        if not os.path.exists(os.path.join(path_output_in, basename_in + '_visualizations', metric)):
            os.makedirs(os.path.join(path_output_in, basename_in + '_visualizations', metric))

        for position in positions_in:
            pattern = re.compile(r'([A-H]\d{2})')
            pos2 = pattern.search(position)
            if pos2 is not None:
                pos2 = pos2.group()

            condition_info = assigned_df_in[assigned_df_in['Well'] == pos2]
            if condition_info.empty:
                print(f"Condition info is empty for {position}")
                continue  # Skip if no matching condition

            condition_name = condition_info['Condition_Name'].values[0]

            if condition_name not in grouped_data:
                grouped_data[condition_name] = []

            grouped_data[condition_name].append(df[position])

        # Plot data for each condition
        plt.figure(figsize=(10, 6))
        for condition_name, data_list in grouped_data.items():
            data_array = pd.concat(data_list, axis=1)
            mean = data_array.mean(axis=1)
            std = data_array.std(axis=1)

            plt.plot(time_hours, mean, label=condition_name)
            plt.fill_between(time_hours, mean - std, mean + std, alpha=0.2)

        plt.title(f'{metric} vs Time')
        plt.xlabel('Time (Hours)')
        plt.ylabel(metric)
        plt.legend()
        plt.grid(True)


        # Save the plot
        output_filename =os.path.join(path_output_in, basename_in + '_' + metric + '.png')
        plt.savefig(output_filename)
        plt.close()
        print(f"Plots saved in ", output_filename)